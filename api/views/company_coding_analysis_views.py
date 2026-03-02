from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from api.services.leetcode_service import LeetCodeService
from api.services.coding_profile_analysis_service import CodingProfileAnalysisService
from api.coding_platform_models import LeetCodeAnalysisHistory
from urllib.parse import urlparse
import csv
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime

logger = logging.getLogger(__name__)


def extract_leetcode_urls_from_text(text):
    """
    Extracts LeetCode URLs from text.
    Handles formats like:
    - https://leetcode.com/u/username/
    - https://leetcode.com/username/
    - leetcode.com/u/username
    """
    import re
    urls = []
    # Pattern to match LeetCode URLs
    pattern = r'(?:https?://)?(?:www\.)?leetcode\.com/(?:u/)?([^/\s]+)'
    matches = re.findall(pattern, text, re.IGNORECASE)
    
    for match in matches:
        username = match.strip('/')
        if username:
            url = f"https://leetcode.com/u/{username}/"
            if url not in urls:
                urls.append(url)
    
    return urls


def extract_username_from_leetcode_url(url):
    """Extract username from LeetCode URL."""
    try:
        parsed = urlparse(url)
        path = parsed.path.strip('/')
        parts = path.split('/')
        if len(parts) >= 2 and parts[0] == 'u':
            return parts[1]
        elif len(parts) >= 1:
            return parts[-1]
        return None
    except Exception:
        return None


class CompanyLeetCodeAnalysisView(APIView):
    """
    API endpoint for companies to analyze multiple LeetCode profiles.
    Supports:
    - Single URL input
    - Multiple URLs (comma-separated or newline-separated)
    - CSV file upload with LeetCode URLs
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Process LeetCode URLs and return analysis results.
        
        Expected input:
        - urls: string (comma or newline separated URLs)
        - csv_file: file (CSV file containing LeetCode URLs)
        """
        try:
            leetcode_urls = []
            target_role = request.data.get('target_role', 'Mid-Level')
            
            # Handle CSV file upload
            csv_name_mapping = {}  # Map URL to name from CSV
            if 'csv_file' in request.FILES:
                csv_file = request.FILES['csv_file']
                try:
                    # Read CSV file
                    decoded_file = csv_file.read().decode('utf-8')
                    csv_reader = csv.reader(io.StringIO(decoded_file))
                    
                    # Extract URLs from all cells and track names
                    for row_idx, row in enumerate(csv_reader):
                        # Assume first column might be name, rest might have URLs
                        row_name = row[0] if row else f"Row {row_idx + 1}"
                        
                        for cell_idx, cell in enumerate(row):
                            urls = extract_leetcode_urls_from_text(cell)
                            for url in urls:
                                leetcode_urls.append(url)
                                # Map URL to name (use first non-empty cell as name, or row number)
                                if url not in csv_name_mapping:
                                    # Try to find a name in the row (first cell that looks like a name)
                                    name_candidate = row[0] if row[0] and not any(c in row[0].lower() for c in ['http', 'leetcode', 'www', '@']) else f"Row {row_idx + 1}"
                                    csv_name_mapping[url] = name_candidate
                except Exception as e:
                    logger.error(f"Error parsing CSV file: {str(e)}")
                    return Response(
                        {"error": f"Failed to parse CSV file: {str(e)}"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Handle direct URL input
            if 'urls' in request.data:
                urls_input = request.data.get('urls', '')
                if urls_input:
                    # Split by comma or newline
                    url_list = urls_input.replace('\n', ',').split(',')
                    for url in url_list:
                        url = url.strip()
                        if url:
                            # Extract URLs from the text (handles both plain URLs and text with URLs)
                            extracted = extract_leetcode_urls_from_text(url)
                            if extracted:
                                leetcode_urls.extend(extracted)
                            elif url.startswith('http') and 'leetcode.com' in url:
                                leetcode_urls.append(url)
            
            if not leetcode_urls:
                return Response(
                    {"error": "No LeetCode URLs found. Please provide URLs or upload a CSV file."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Remove duplicates while preserving order
            seen = set()
            unique_urls = []
            for url in leetcode_urls:
                if url not in seen:
                    seen.add(url)
                    unique_urls.append(url)
            
            # Limit to 100 URLs per request
            if len(unique_urls) > 100:
                return Response(
                    {"error": f"Too many URLs. Maximum 100 URLs allowed. Found {len(unique_urls)}."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Fetch stats for each URL (with rate limiting)
            results = []
            import time
            for idx, url in enumerate(unique_urls):
                # Add delay to avoid rate limiting
                if idx > 0:
                    # Progressive delay: 0.5s between requests, 2s every 10 requests
                    if idx % 10 == 0:
                        time.sleep(2)  # Longer delay every 10 requests
                    else:
                        time.sleep(0.5)  # Small delay between requests
                username = extract_username_from_leetcode_url(url)
                csv_name = csv_name_mapping.get(url, None)  # Get name from CSV if available
                
                if not username:
                    results.append({
                        "url": url,
                        "username": None,
                        "csv_name": csv_name,
                        "error": "Invalid LeetCode URL format",
                        "stats": None,
                        "analysis": None,
                        "analyzed_at": timezone.now().isoformat()
                    })
                    continue
                
                try:
                    # Fetch stats
                    stats_result = LeetCodeService.get_user_stats(username, target_role)
                    
                    # Check if result is an error dict or actual stats
                    if stats_result and isinstance(stats_result, dict):
                        # Check if it's a pure error (no stats data)
                        if 'error' in stats_result and 'total_solved' not in stats_result:
                            # This is a pure error response
                            error_msg = stats_result.get('error', 'Failed to fetch stats')
                            results.append({
                                "url": url,
                                "username": username,
                                "csv_name": csv_name,
                                "error": error_msg,
                                "stats": None,
                                "analysis": None,
                                "analyzed_at": timezone.now().isoformat()
                            })
                        elif 'total_solved' in stats_result or 'ranking' in stats_result:
                            # This is actual stats data (even if some fields are zero)
                            stats = stats_result.copy()
                            # Remove error field if present (we have stats, so it's not a complete failure)
                            stats.pop('error', None)
                            stats.pop('exists', None)
                            
                            # Create a temporary profile-like object for analysis
                            class TempProfile:
                                def __init__(self, username, stats):
                                    self.username = username
                                    self.platform = 'leetcode'
                                    self.stats = stats
                            
                            temp_profile = TempProfile(username, stats)
                            
                            # Get AI analysis
                            analysis = None
                            try:
                                analysis = CodingProfileAnalysisService.analyze_profile(temp_profile)
                            except Exception as e:
                                logger.warning(f"Failed to generate analysis for {username}: {str(e)}")
                            
                            # Save to history for progress tracking
                            try:
                                LeetCodeAnalysisHistory.objects.create(
                                    company=request.user,
                                    employee_identifier=csv_name or username,
                                    leetcode_username=username,
                                    leetcode_url=url,
                                    total_solved=stats.get('total_solved', 0),
                                    easy_solved=stats.get('easy_solved', 0),
                                    medium_solved=stats.get('medium_solved', 0),
                                    hard_solved=stats.get('hard_solved', 0),
                                    problem_solving_score=stats.get('problem_solving_score', 0),
                                    ranking=stats.get('ranking', 0),
                                    acceptance_rate=stats.get('acceptance_rate', 0),
                                    current_streak=stats.get('current_streak', 0),
                                    max_streak=stats.get('max_streak', 0),
                                    activity_status=stats.get('activity_status', 'Unknown'),
                                    full_stats=stats,
                                    analysis_data=analysis or {},
                                )
                            except Exception as e:
                                logger.warning(f"Failed to save history for {username}: {str(e)}")
                            
                            results.append({
                                "url": url,
                                "username": username,
                                "csv_name": csv_name,
                                "error": None,
                                "stats": stats,
                                "analysis": analysis,
                                "analyzed_at": timezone.now().isoformat()
                            })
                    else:
                        # No stats returned (None or unexpected format)
                        results.append({
                            "url": url,
                            "username": username,
                            "csv_name": csv_name,
                            "error": "Failed to fetch stats. User may not exist or profile is private.",
                            "stats": None,
                            "analysis": None,
                            "analyzed_at": timezone.now().isoformat()
                        })
                except Exception as e:
                    logger.error(f"Error processing {url}: {str(e)}")
                    results.append({
                        "url": url,
                        "username": username,
                        "csv_name": csv_name,
                        "error": str(e),
                        "stats": None,
                        "analysis": None,
                        "analyzed_at": timezone.now().isoformat()
                    })
            
            return Response({
                "total_urls": len(unique_urls),
                "successful": len([r for r in results if r.get('stats')]),
                "failed": len([r for r in results if not r.get('stats')]),
                "results": results
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Unexpected error in CompanyLeetCodeAnalysisView: {str(e)}", exc_info=True)
            return Response(
                {"error": f"An unexpected error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CompanyLeetCodeExportView(APIView):
    """
    Export LeetCode analysis results to Excel or CSV.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        """
        Export results to Excel or CSV.
        
        Expected input:
        - results: list of result objects (same format as CompanyLeetCodeAnalysisView response)
        - format: 'excel' or 'csv' (default: 'excel')
        - filter_month: string (optional)
        - filter_year: string (optional)
        """
        try:
            results = request.data.get('results', [])
            export_format = request.data.get('format', 'excel').lower()
            filter_month = request.data.get('filter_month', 'all')
            filter_year = request.data.get('filter_year', 'all')
            
            if not results:
                return Response(
                    {"error": "No results to export."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if export_format == 'csv':
                return self._export_csv(results, filter_month, filter_year)
            else:
                return self._export_excel(results, filter_month, filter_year)
                
        except Exception as e:
            logger.error(f"Error exporting results: {str(e)}", exc_info=True)
            return Response(
                {"error": f"Export failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    def _calculate_solved_count(self, stats, filter_month, filter_year):
        """Calculate solved count based on filters."""
        if not stats:
            return 0
            
        default_solved = stats.get('total_solved', 0)
        
        if filter_month == 'all' and filter_year == 'all':
            return default_solved
            
        submission_calendar = stats.get('submission_calendar')
        if not submission_calendar:
            return default_solved
            
        try:
            target_month = int(filter_month) if filter_month != 'all' else None
            target_year = int(filter_year) if filter_year != 'all' else None
            count = 0
            
            for ts, val in submission_calendar.items():
                try:
                    dt = datetime.fromtimestamp(int(ts))
                    month_match = True
                    year_match = True
                    
                    if target_month is not None:
                        month_match = dt.month == target_month
                        
                    if target_year is not None:
                        year_match = dt.year == target_year
                        
                    if month_match and year_match:
                        count += val
                except:
                    continue
                    
            return count
        except:
            return default_solved

    def _get_solved_header(self, filter_month, filter_year):
        """Get header for solved column."""
        if filter_month == 'all' and filter_year == 'all':
            return 'Total Solved'
        
        parts = []
        if filter_month != 'all':
            try:
                month_name = datetime(2000, int(filter_month), 1).strftime('%b')
                parts.append(month_name)
            except:
                parts.append(filter_month)
                
        if filter_year != 'all':
            parts.append(str(filter_year))
            
        return f"Solved ({' '.join(parts)})"
    
    def _export_csv(self, results, filter_month, filter_year):
        """Export results to CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="leetcode_analysis.csv"'
        
        writer = csv.writer(response)
        
        solved_header = self._get_solved_header(filter_month, filter_year)
        
        # Header row
        writer.writerow([
            'CSV Name', 'URL', 'Username', 'Real Name', 'Avatar', 'Company', 'School', 'Country',
            solved_header, 'Easy', 'Medium', 'Hard',
            'Total Submissions', 'Acceptance Rate %', 'Weighted Acceptance Rate %',
            'Problem Solving Score', 'Current Streak', 'Max Streak', 'Avg Weekly Subs', 'Activity Status',
            'Ranking', 'Reputation', 'Engagement Level', 'Star Rating',
            'Contest Rating', 'Contests Attended', 'Contest Ranking', 'Contest Top %',
            'Recent Contests', 'Recent Avg Rating',
            'Estimated Level', 'Summary', 'Strengths', 'Recommendations', 'Insights', 'Analyzed At', 'Error'
        ])
        
        # Data rows
        for result in results:
            stats = result.get('stats', {}) or {}
            analysis = result.get('analysis', {}) or {}
            
            strengths = ', '.join(analysis.get('strengths', [])) if analysis.get('strengths') else ''
            recommendations = ', '.join(analysis.get('recommendations', [])) if analysis.get('recommendations') else ''
            insights = ', '.join(analysis.get('insights', [])) if analysis.get('insights') else ''
            
            analyzed_at = result.get('analyzed_at', '')
            if analyzed_at:
                try:
                    dt = datetime.fromisoformat(analyzed_at.replace('Z', '+00:00'))
                    analyzed_at = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            
            solved_count = self._calculate_solved_count(stats, filter_month, filter_year)
            
            writer.writerow([
                result.get('csv_name', ''),
                result.get('url', ''),
                result.get('username', ''),
                stats.get('real_name', ''),
                stats.get('avatar', ''),
                stats.get('company', ''),
                stats.get('school', ''),
                stats.get('country', ''),
                solved_count,
                stats.get('easy_solved', 0),
                stats.get('medium_solved', 0),
                stats.get('hard_solved', 0),
                stats.get('total_submissions', 0),
                stats.get('acceptance_rate', 0),
                stats.get('weighted_acceptance_rate', 0),
                stats.get('problem_solving_score', 0),
                stats.get('current_streak', 0),
                stats.get('max_streak', 0),
                stats.get('avg_weekly_submissions', 0),
                stats.get('activity_status', 'Unknown'),
                stats.get('ranking', 0),
                stats.get('reputation', 0),
                stats.get('community_engagement', 'Low'),
                stats.get('star_rating', 0),
                stats.get('contest_rating', 0),
                stats.get('contests_attended', 0),
                stats.get('contest_global_ranking', 0),
                stats.get('contest_top_percentage', 0),
                stats.get('recent_contests', 0),
                stats.get('recent_avg_rating', 0),
                analysis.get('estimated_level', ''),
                analysis.get('summary', ''),
                strengths,
                recommendations,
                insights,
                analyzed_at,
                result.get('error', '')
            ])
        
        return response
    
    def _export_excel(self, results, filter_month, filter_year):
        """Export results to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "LeetCode Analysis"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        solved_header = self._get_solved_header(filter_month, filter_year)
        
        # Headers
        headers = [
            'CSV Name', 'URL', 'Username', 'Real Name', 'Avatar', 'Company', 'School', 'Country',
            solved_header, 'Easy', 'Medium', 'Hard',
            'Total Submissions', 'Acceptance Rate %', 'Weighted Acceptance Rate %',
            'Problem Solving Score', 'Current Streak', 'Max Streak', 'Avg Weekly Subs', 'Activity Status',
            'Ranking', 'Reputation', 'Engagement Level', 'Star Rating',
            'Contest Rating', 'Contests Attended', 'Contest Ranking', 'Contest Top %',
            'Recent Contests', 'Recent Avg Rating',
            'Estimated Level', 'Summary', 'Strengths', 'Recommendations', 'Insights', 'Analyzed At', 'Error'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        for row_num, result in enumerate(results, 2):
            stats = result.get('stats', {}) or {}
            analysis = result.get('analysis', {}) or {}
            
            strengths = ', '.join(analysis.get('strengths', [])) if analysis.get('strengths') else ''
            recommendations = ', '.join(analysis.get('recommendations', [])) if analysis.get('recommendations') else ''
            insights = ', '.join(analysis.get('insights', [])) if analysis.get('insights') else ''
            
            solved_count = self._calculate_solved_count(stats, filter_month, filter_year)
            
            ws.cell(row=row_num, column=1, value=result.get('csv_name', ''))
            ws.cell(row=row_num, column=2, value=result.get('url', ''))
            ws.cell(row=row_num, column=3, value=result.get('username', ''))
            ws.cell(row=row_num, column=4, value=stats.get('real_name', ''))
            ws.cell(row=row_num, column=5, value=stats.get('avatar', ''))
            ws.cell(row=row_num, column=6, value=stats.get('company', ''))
            ws.cell(row=row_num, column=7, value=stats.get('school', ''))
            ws.cell(row=row_num, column=8, value=stats.get('country', ''))
            ws.cell(row=row_num, column=9, value=solved_count)
            ws.cell(row=row_num, column=10, value=stats.get('easy_solved', 0))
            ws.cell(row=row_num, column=11, value=stats.get('medium_solved', 0))
            ws.cell(row=row_num, column=12, value=stats.get('hard_solved', 0))
            ws.cell(row=row_num, column=13, value=stats.get('total_submissions', 0))
            ws.cell(row=row_num, column=14, value=stats.get('acceptance_rate', 0))
            ws.cell(row=row_num, column=15, value=stats.get('weighted_acceptance_rate', 0))
            ws.cell(row=row_num, column=16, value=stats.get('problem_solving_score', 0))
            ws.cell(row=row_num, column=17, value=stats.get('current_streak', 0))
            ws.cell(row=row_num, column=18, value=stats.get('max_streak', 0))
            ws.cell(row=row_num, column=19, value=stats.get('avg_weekly_submissions', 0))
            ws.cell(row=row_num, column=20, value=stats.get('activity_status', 'Unknown'))
            ws.cell(row=row_num, column=21, value=stats.get('ranking', 0))
            ws.cell(row=row_num, column=22, value=stats.get('reputation', 0))
            ws.cell(row=row_num, column=23, value=stats.get('community_engagement', 'Low'))
            ws.cell(row=row_num, column=24, value=stats.get('star_rating', 0))
            ws.cell(row=row_num, column=25, value=stats.get('contest_rating', 0))
            ws.cell(row=row_num, column=26, value=stats.get('contests_attended', 0))
            ws.cell(row=row_num, column=27, value=stats.get('contest_global_ranking', 0))
            ws.cell(row=row_num, column=28, value=stats.get('contest_top_percentage', 0))
            ws.cell(row=row_num, column=29, value=stats.get('recent_contests', 0))
            ws.cell(row=row_num, column=30, value=stats.get('recent_avg_rating', 0))
            ws.cell(row=row_num, column=31, value=analysis.get('estimated_level', ''))
            ws.cell(row=row_num, column=32, value=analysis.get('summary', ''))
            ws.cell(row=row_num, column=33, value=strengths)
            ws.cell(row=row_num, column=34, value=recommendations)
            ws.cell(row=row_num, column=35, value=insights)
            
            # Analyzed At
            analyzed_at = result.get('analyzed_at', '')
            if analyzed_at:
                try:
                    dt = datetime.fromisoformat(analyzed_at.replace('Z', '+00:00'))
                    analyzed_at = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            ws.cell(row=row_num, column=36, value=analyzed_at)
            ws.cell(row=row_num, column=37, value=result.get('error', ''))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="leetcode_analysis.xlsx"'
        wb.save(response)
        
        return response

